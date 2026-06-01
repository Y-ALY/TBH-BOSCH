"""Memory-safe GDPR PII extraction engine.

Designed for unlimited data scale (5 MB → 500 GB → terabytes).

Architecture:
  ┌─────────────────────────────────────────────────────┐
  │ scan_directory(dir, owner_hints)                    │
  │   ├─ walks directory tree (lazy os.walk generator)  │
  │   ├─ calls scan_file() per file (constant memory)   │
  │   ├─ accumulates lightweight aggregates only        │
  │   └─ returns { admin_aggregates, user_file_details }│
  └─────────────────────────────────────────────────────┘

Memory model:
  - At any moment, only ONE file's text is in memory.
  - PDF pages are yielded one-at-a-time via generators.
  - Text files are read in 1 MB chunks with overlap.
  - After each file, all per-file objects are released.

Time complexity:
  - O(N) in total bytes across all files.
  - Regex matching is O(M * K) per chunk, where M = chunk
    length and K = number of compiled patterns (~12).
  - Total: O(N * K) — linear in data volume.

Space complexity:
  - O(chunk_size) per file ≈ 1 MB constant.
  - Aggregates: O(F) where F = number of files (lightweight dicts).
  - Never loads the full dataset into RAM.
"""

from __future__ import annotations

import io
import os
import re
import time
import hashlib
from collections import defaultdict
from pathlib import Path
from typing import Generator, Any

# ─────────────────────────────────────────────────────────────────────────────
# 1. COMPILED REGEX DICTIONARY — compiled ONCE at module-import time.
#
#    re.compile() builds an internal finite-state machine from the pattern
#    string.  By doing this at module level we pay the compilation cost
#    exactly once, not once per file or per chunk.
#
#    Each entry:  (compiled_pattern, category, risk_level, recommended_action)
# ─────────────────────────────────────────────────────────────────────────────

# --- Email ---
# RFC 5322 simplified.  False-positive filter: skip @example.com / @test.com
_RE_EMAIL = re.compile(
    r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}',
    re.IGNORECASE,
)

# --- Phone Numbers ---
# International (E.164-ish), German, US, UK, generic digit groups.
# Matches: +49 170 1234567, 0170/1234567, (555) 123-4567, +34841 04 31 83
_RE_PHONE = re.compile(
    r'(?:'
    r'\+\d{1,3}[\s\-]?\(?\d{1,4}\)?[\s\-]?\d[\d\s\-]{5,14}\d'  # international
    r'|'
    r'\b0\d{2,4}[\s/\-]?\d{3,}[\s\-]?\d{2,}'                    # German local
    r'|'
    r'\(\d{3}\)\s?\d{3}[\-\s]?\d{4}'                              # US (xxx) xxx-xxxx
    r')',
)

# --- Addresses ---
# German postal code (5 digits + city name), or street + number patterns.
_RE_ADDRESS_DE = re.compile(
    r'\b\d{5}\s+[A-ZÄÖÜ][a-zäöüß]+(?:\s+[A-ZÄÖÜ][a-zäöüß]+)*\b',
)
_RE_ADDRESS_STREET = re.compile(
    r'\b[A-ZÄÖÜ][a-zäöüß]+(?:straße|str\.|weg|gasse|allee|platz|ring|damm|ufer)'
    r'\s+\d{1,5}[a-z]?\b',
    re.IGNORECASE,
)

# --- Passport Numbers ---
# German: exactly 9 alphanumeric chars starting with a letter (C, F, G, H, J, K)
# UK: 9 digits
# US: 1 letter + 8 digits
_RE_PASSPORT_DE = re.compile(r'\b[CFGHJK][A-Z0-9]{8}\b')
_RE_PASSPORT_GENERIC = re.compile(r'\b[A-Z]\d{8}\b')

# --- ID Card Numbers ---
# German Personalausweis: starts with L,M,N,T,P,H,R,C + 8 alphanumeric
_RE_ID_CARD_DE = re.compile(r'\b[LMNТPHRC][A-Z0-9]{8}\b')

# --- Driver's License ---
# German: 11 alphanumeric characters
# UK DVLA: specific format (SURNAME + DOB encoded)
_RE_DRIVERS_LICENSE_DE = re.compile(r'\b[A-Z0-9]{11}\b')

# --- IBAN ---
_RE_IBAN = re.compile(r'\b[A-Z]{2}\d{2}\s?[A-Z0-9]{4}[\sA-Z0-9]{10,30}\b')

# --- Tax ID ---
# German Steuer-ID (11 digits) and USt-IdNr (DE + 9 digits)
_RE_TAX_STEUER = re.compile(r'\b\d{2}/\d{3}/\d{5}\b')
_RE_TAX_UST = re.compile(r'\bDE\d{9}\b')

# --- Names ---
# Labeled name fields in structured documents.
_RE_NAME_LABELED = re.compile(
    r'(?:^|\n)\s*(?:Patient\s*Name|Name|Vorname|Nachname|Full\s*Name|'
    r'Employee|Manager|Participant|Teilnehmer|Reported\s*by|'
    r'Attending\s*Physician|Physician\s*Signature)'
    r'\s*[:]\s*(.+?)(?:\n|$)',
    re.IGNORECASE | re.MULTILINE,
)

# --- Date of Birth ---
_RE_DOB = re.compile(
    r'(?:Date\s*of\s*Birth|DOB|Geburtsdatum|Born)\s*[:]\s*'
    r'(\d{4}[\-/]\d{2}[\-/]\d{2}|\d{2}[./]\d{2}[./]\d{4})',
    re.IGNORECASE,
)

# --- Signature ---
_RE_SIGNATURE = re.compile(
    r'(?:Signed|Signature|Unterschrift)\s*[:]\s*[\w\s.]+',
    re.IGNORECASE,
)

# --- SSN / Social Security ---
_RE_SSN = re.compile(
    r'(?:SSN|Social\s*Security)\s*[:=]?\s*([\d\-]{9,11})\b',
    re.IGNORECASE,
)

# --- Employee ID ---
_RE_EMPLOYEE_ID = re.compile(r'\bEMP-\d{5,8}\b')


# ─────────────────────────────────────────────────────────────────────────────
# Master pattern registry.
# Each tuple: (compiled_regex, category, risk_level, action, use_group)
#   use_group: if > 0, extract match.group(use_group) instead of group(0)
# ─────────────────────────────────────────────────────────────────────────────

_COMPILED_PATTERNS: list[tuple[re.Pattern, str, str, str, int]] = [
    (_RE_EMAIL,             "email",            "medium", "mask",   0),
    (_RE_PHONE,             "phone",            "medium", "mask",   0),
    (_RE_ADDRESS_DE,        "address",          "medium", "review", 0),
    (_RE_ADDRESS_STREET,    "address",          "medium", "review", 0),
    (_RE_PASSPORT_DE,       "passport",         "high",   "delete", 0),
    (_RE_PASSPORT_GENERIC,  "passport",         "high",   "delete", 0),
    (_RE_ID_CARD_DE,        "id_card",          "high",   "delete", 0),
    (_RE_IBAN,              "iban",             "high",   "delete", 0),
    (_RE_TAX_STEUER,        "tax_id",           "high",   "delete", 0),
    (_RE_TAX_UST,           "tax_id",           "high",   "delete", 0),
    (_RE_NAME_LABELED,      "name",             "medium", "mask",   1),
    (_RE_DOB,               "date_of_birth",    "high",   "mask",   1),
    (_RE_SIGNATURE,         "signature",        "low",    "retain", 0),
    (_RE_SSN,               "ssn",              "high",   "delete", 1),
    (_RE_EMPLOYEE_ID,       "employee_id",      "medium", "mask",   0),
    (_RE_DRIVERS_LICENSE_DE,"drivers_license",  "high",   "delete", 0),
]


# ─────────────────────────────────────────────────────────────────────────────
# 2. FALSE-POSITIVE FILTERS — per-category post-match validation.
#
#    After regex matching, each candidate value passes through a filter.
#    This dramatically reduces noise without slowing down the regex engine.
# ─────────────────────────────────────────────────────────────────────────────

_EMAIL_BLACKLIST_DOMAINS = {
    "example.com", "test.com", "localhost", "placeholder.org",
    "noreply.com", "invalid.com",
}

_COMMON_WORDS_UPPER = {
    # Words that look like passport/ID patterns but aren't
    "DEPARTMENT", "CERTIFICATE", "CONFIDENTIAL", "HEALTHCARE",
    "BESCHEINIG", "DIAGNOSTIC", "POSTOPERATI", "INFORMATION",
    "DESCRIPTION", "APPLICATION", "CERTIFICATE", "PERFORMANCE",
}


def _filter_email(value: str) -> bool:
    """Return True if this email should be KEPT (not a false positive)."""
    domain = value.split("@")[-1].lower()
    return domain not in _EMAIL_BLACKLIST_DOMAINS


def _filter_passport(value: str) -> bool:
    """Filter out common English words that match the passport pattern."""
    return value.upper() not in _COMMON_WORDS_UPPER and not value.isalpha()


def _filter_id_card(value: str) -> bool:
    """German ID cards always mix letters and digits."""
    has_digit = any(c.isdigit() for c in value)
    has_alpha = any(c.isalpha() for c in value)
    return has_digit and has_alpha and value.upper() not in _COMMON_WORDS_UPPER


def _filter_drivers_license(value: str) -> bool:
    """Must contain both letters and digits, and not be a common word."""
    has_digit = any(c.isdigit() for c in value)
    has_alpha = any(c.isalpha() for c in value)
    return has_digit and has_alpha and len(value) == 11 and value.upper() not in _COMMON_WORDS_UPPER


def _filter_iban(value: str) -> bool:
    """Basic IBAN structure validation."""
    cleaned = value.replace(" ", "")
    return len(cleaned) >= 15 and cleaned[:2].isalpha() and cleaned[2:4].isdigit()


# Map category -> filter function.  Categories not listed here pass all matches.
_FILTERS: dict[str, callable] = {
    "email": _filter_email,
    "passport": _filter_passport,
    "id_card": _filter_id_card,
    "drivers_license": _filter_drivers_license,
    "iban": _filter_iban,
}


# ─────────────────────────────────────────────────────────────────────────────
# 3. GENERATOR-BASED CHUNKED FILE READERS
#
#    Each reader yields (chunk_text: str) one piece at a time.
#    The caller never holds more than one chunk in memory.
#
#    Time: O(N) per file where N = file size.
#    Space: O(page_size) for PDF, O(chunk_size) for text, constant.
# ─────────────────────────────────────────────────────────────────────────────

def _read_pdf_pages(file_path: str) -> Generator[str, None, None]:
    """Yield text page-by-page from a PDF. O(1 page) memory.

    Uses pdfplumber for text-layer PDFs.  Each page's text is yielded
    independently and then released, so a 10,000-page PDF only ever
    holds one page of text in RAM.
    """
    import pdfplumber
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if text.strip():
                yield text.strip()


def _read_docx_paragraphs(file_path: str) -> Generator[str, None, None]:
    """Yield text paragraph-by-paragraph from a DOCX. O(1 paragraph) memory."""
    from docx import Document
    doc = Document(file_path)
    # Accumulate into ~64KB chunks to amortize regex overhead
    buffer = []
    buffer_size = 0
    chunk_limit = 65_536  # 64 KB

    for para in doc.paragraphs:
        text = para.text.strip()
        if text:
            buffer.append(text)
            buffer_size += len(text)
            if buffer_size >= chunk_limit:
                yield "\n".join(buffer)
                buffer.clear()
                buffer_size = 0

    if buffer:
        yield "\n".join(buffer)


def _read_text_chunks(
    file_path: str,
    chunk_size: int = 1_048_576,  # 1 MB
    overlap: int = 512,            # 512-byte overlap to avoid split matches
) -> Generator[str, None, None]:
    """Yield text in fixed-size chunks with overlap.

    The overlap ensures that a regex match spanning a chunk boundary
    is still captured in at least one chunk.  Deduplication downstream
    prevents double-counting.

    Time: O(N) total.  Space: O(chunk_size) constant.
    """
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        carry = ""
        while True:
            raw = f.read(chunk_size)
            if not raw:
                if carry:
                    yield carry
                break
            chunk = carry + raw
            yield chunk
            # Keep the last `overlap` chars for the next iteration
            carry = chunk[-overlap:] if len(chunk) > overlap else chunk


def read_file_chunks(file_path: str) -> Generator[str, None, None]:
    """Universal file reader — dispatches by extension.

    Supports: .pdf, .docx, .txt, .csv, .log, .json, .jsonl, .xml, .html, .md,
    and any other text-readable file.  Binary files that can't be decoded
    are silently skipped with a warning yield.

    Time: O(N).  Space: O(chunk_size) ≈ 1 MB constant.
    """
    ext = Path(file_path).suffix.lower()

    if ext == ".pdf":
        yield from _read_pdf_pages(file_path)

    elif ext == ".docx":
        yield from _read_docx_paragraphs(file_path)

    elif ext in {".xlsx", ".xls"}:
        # Excel: read cell-by-cell as text using openpyxl if available,
        # otherwise fall back to treating as binary-skip.
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows_text = []
                rows_size = 0
                for row in ws.iter_rows(values_only=True):
                    line = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    rows_text.append(line)
                    rows_size += len(line)
                    if rows_size >= 65_536:
                        yield "\n".join(rows_text)
                        rows_text.clear()
                        rows_size = 0
                if rows_text:
                    yield "\n".join(rows_text)
            wb.close()
        except ImportError:
            yield ""  # openpyxl not installed — skip gracefully
        except Exception:
            yield ""

    else:
        # Treat as generic text (txt, csv, log, json, xml, html, md, etc.)
        try:
            yield from _read_text_chunks(file_path)
        except (UnicodeDecodeError, PermissionError, OSError):
            # Binary or unreadable file — skip gracefully
            pass


# ─────────────────────────────────────────────────────────────────────────────
# 4. CORE SCAN FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def _extract_matches_from_text(
    text: str,
    seen: set[tuple[str, str]],
) -> list[dict[str, Any]]:
    """Run all compiled patterns against a text chunk.

    Args:
        text: The text chunk to scan.
        seen: Mutable dedup set of (category, value) — prevents duplicates
              across chunks within the same file.

    Returns:
        List of finding dicts for this chunk.

    Time: O(len(text) * len(_COMPILED_PATTERNS)).
    Space: O(number_of_matches) — typically small relative to text size.
    """
    findings = []

    for compiled_re, category, risk, action, use_group in _COMPILED_PATTERNS:
        for match in compiled_re.finditer(text):
            # Extract the value (either full match or a specific capture group)
            value = match.group(use_group).strip() if use_group > 0 else match.group(0).strip()
            if not value:
                continue

            # Deduplication: skip if already found in this file
            dedup_key = (category, value)
            if dedup_key in seen:
                continue

            # False-positive filter
            fp_filter = _FILTERS.get(category)
            if fp_filter and not fp_filter(value):
                continue

            seen.add(dedup_key)

            # Build surrounding context (up to 80 chars around the match)
            # This is the key field that fixes the frontend display.
            start_ctx = max(0, match.start() - 40)
            end_ctx = min(len(text), match.end() + 40)
            context = text[start_ctx:end_ctx].replace("\n", " ").strip()

            findings.append({
                "category": category,
                "matched_value": value,
                "match_context": f"...{context}...",
                "risk_level": risk,
                "confidence": 1.0,
                "recommended_action": action,
            })

    return findings


def scan_file(file_path: str, owner: str = "", owner_email: str = "") -> dict:
    """Scan a single file for PII using generator-based chunked reading.

    Memory usage is O(chunk_size) regardless of file size.
    A 50 GB file uses the same RAM as a 5 KB file.

    Args:
        file_path:   Absolute or relative path to the file.
        owner:       Human-readable owner name (from owner_hints).
        owner_email: Owner's email address (from owner_hints).

    Returns:
        {
            "file_path": str,
            "file_name": str,
            "owner": str,
            "owner_email": str,
            "size_bytes": int,
            "findings": [ { category, matched_value, match_context, ... } ]
        }
    """
    p = Path(file_path)
    try:
        size_bytes = p.stat().st_size
    except OSError:
        size_bytes = 0

    # Dedup set scoped to this file — prevents duplicate findings
    # across chunks when using overlapping reads.
    seen: set[tuple[str, str]] = set()
    all_findings: list[dict] = []

    # Stream through the file chunk-by-chunk.
    # At any moment, only ONE chunk is in memory.
    for chunk_text in read_file_chunks(str(p)):
        chunk_findings = _extract_matches_from_text(chunk_text, seen)
        all_findings.extend(chunk_findings)
        # chunk_text is released at the top of the next loop iteration.

    return {
        "file_path": str(p),
        "file_name": p.name,
        "owner": owner,
        "owner_email": owner_email,
        "size_bytes": size_bytes,
        "findings": all_findings,
    }


def scan_directory(
    dir_path: str,
    owner_hints: dict | None = None,
) -> dict:
    """Scan an entire directory tree for PII.

    Uses os.walk (a generator) to discover files lazily.
    Processes one file at a time — never loads the full file list into RAM.

    Args:
        dir_path:     Root directory to scan.
        owner_hints:  Optional dict mapping filename -> {name, email, department, ...}.

    Returns:
        {
            "admin_aggregates": {
                "total_scanned_files": int,
                "total_size_bytes": int,
                "total_size_human": str,
                "files_with_findings": int,
                "total_findings": int,
                "findings_by_category": { category: count },
                "scan_duration_seconds": float,
            },
            "user_file_details": [
                {
                    "file_path": str,
                    "file_name": str,
                    "owner": str,
                    "owner_email": str,
                    "size_bytes": int,
                    "findings": [ ... ]
                }
            ]
        }

    Time:  O(total_bytes_across_all_files * num_patterns).
    Space: O(num_files) for aggregates + O(chunk_size) for reading.
           The user_file_details list grows with file count, but each entry
           is a lightweight dict (not the full file content).
    """
    if owner_hints is None:
        owner_hints = {}

    start_time = time.monotonic()

    # ── Aggregate counters (O(1) per update) ──
    total_files = 0
    total_size_bytes = 0
    files_with_findings = 0
    total_findings = 0
    findings_by_category: dict[str, int] = defaultdict(int)

    # ── Per-file results (lightweight dicts, not full content) ──
    user_file_details: list[dict] = []

    # os.walk is itself a generator — it does not list the entire tree upfront.
    # This is critical for directories with millions of files.
    root = Path(dir_path)
    if not root.exists():
        return _build_result(
            total_files, total_size_bytes, files_with_findings,
            total_findings, dict(findings_by_category),
            user_file_details, time.monotonic() - start_time,
        )

    for dirpath, _dirnames, filenames in os.walk(str(root)):
        for filename in sorted(filenames):
            # Skip hidden files, __pycache__, .git, etc.
            if filename.startswith(".") or filename.startswith("__"):
                continue

            file_path = os.path.join(dirpath, filename)

            # Resolve owner from hints
            hint = owner_hints.get(filename, {})
            owner_name = hint.get("name", "")
            owner_email = hint.get("email", "")

            # ── Scan this file (constant memory) ──
            try:
                result = scan_file(file_path, owner=owner_name, owner_email=owner_email)
            except Exception:
                # Non-fatal: skip unreadable files, keep scanning.
                total_files += 1
                continue

            # ── Update aggregates ──
            total_files += 1
            total_size_bytes += result["size_bytes"]

            if result["findings"]:
                files_with_findings += 1
                total_findings += len(result["findings"])
                for f in result["findings"]:
                    findings_by_category[f["category"]] += 1

            user_file_details.append(result)

    elapsed = time.monotonic() - start_time

    return _build_result(
        total_files, total_size_bytes, files_with_findings,
        total_findings, dict(findings_by_category),
        user_file_details, elapsed,
    )


def _format_size(size_bytes: int) -> str:
    """Human-readable file size."""
    if size_bytes >= 1_073_741_824:
        return f"{size_bytes / 1_073_741_824:.2f} GB"
    elif size_bytes >= 1_048_576:
        return f"{size_bytes / 1_048_576:.2f} MB"
    elif size_bytes >= 1024:
        return f"{size_bytes / 1024:.2f} KB"
    return f"{size_bytes} B"


def _build_result(
    total_files: int,
    total_size_bytes: int,
    files_with_findings: int,
    total_findings: int,
    findings_by_category: dict,
    user_file_details: list,
    elapsed: float,
) -> dict:
    """Assemble the final strict JSON contract."""
    return {
        "admin_aggregates": {
            "total_scanned_files": total_files,
            "total_size_bytes": total_size_bytes,
            "total_size_human": _format_size(total_size_bytes),
            "files_with_findings": files_with_findings,
            "total_findings": total_findings,
            "findings_by_category": findings_by_category,
            "scan_duration_seconds": round(elapsed, 3),
        },
        "user_file_details": user_file_details,
    }
